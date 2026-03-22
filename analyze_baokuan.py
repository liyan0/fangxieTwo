# -*- coding: utf-8 -*-
import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')
wb = openpyxl.load_workbook(r'D:\A百家号带货视频\带货文案\爆款参考文案\爆款素材库.xlsx')
ws = wb.active
texts = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    text = str(row[2]) if row[2] else ''
    ltype = str(row[0]) if row[0] else ''
    if text and text not in ('None', ''):
        texts.append((ltype, text))
print(f'共{len(texts)}篇')

# 统计类型
type_count = {}
for t, _ in texts:
    type_count[t] = type_count.get(t, 0) + 1
print('\n=== 类型分布 ===')
for k, v in type_count.items():
    print(f'{k}: {v}篇')

# 分析开头风格
print('\n=== 全部文案开头200字 ===')
for i, (t, text) in enumerate(texts):
    print(f'\n【{i+1}】{t}')
    print(text[:200])
    print('---')
